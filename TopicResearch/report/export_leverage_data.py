"""
导出三国杠杆年度数据（2005-2025）到 Excel
数据来源：ifind EDB
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = os.path.expanduser("~/llmwikify/TopicResearch/report")
OUT_PATH = os.path.join(BASE_DIR, "leverage_data_2005_2025.xlsx")

# 数据定义（2005-2025）
YEARS = list(range(2005, 2026))

# 美国数据
US_GOV_DEBT = [61.6, 62.0, 62.6, 67.5, 82.1, 90.9, 95.5, 99.6, 100.2, 102.1, 99.8, 104.8, 104.3, 105.0, 106.5, 126.1, 123.2, 120.6, 120.4, 121.7, 123.3]
US_HH_LEVERAGE = [92.7, 96.9, 98.7, 96.1, 97.0, 91.7, 87.8, 83.6, 81.8, 79.0, 77.3, 77.4, 77.0, 75.0, 74.6, 77.9, 76.7, 74.5, 71.8, 69.3, 68.1]
US_CORP_LEVERAGE = [64.8, 67.2, 72.1, 74.6, 72.6, 69.0, 68.3, 69.1, 69.9, 71.7, 73.7, 75.2, 77.3, 78.2, 78.3, 86.3, 82.8, 78.9, 75.3, 73.3, 72.2]

# 中国数据
CN_GOV_DEBT = [25.9, 25.2, 28.7, 26.7, 34.0, 33.3, 33.2, 33.8, 36.4, 39.3, 40.8, 49.7, 53.9, 55.6, 59.4, 69.0, 70.1, 77.3, 84.1, 90.4, 99.2]
CN_HH_LEVERAGE = [16.6, 17.2, 18.5, 17.6, 23.1, 26.9, 27.5, 29.5, 32.9, 35.3, 38.5, 43.8, 47.8, 51.2, 55.0, 61.1, 60.6, 60.7, 61.9, 61.4, 59.4]
CN_CORP_LEVERAGE = [98.4, 97.6, 94.6, 93.7, 113.6, 118.4, 116.1, 126.0, 133.5, 139.8, 148.5, 154.4, 153.8, 147.9, 149.0, 159.2, 150.9, 157.7, 163.9, 168.4, 174.6]

# 日本数据
JP_GOV_DEBT = [174.6, 174.1, 173.0, 180.9, 198.8, 205.9, 219.2, 226.1, 229.5, 233.3, 228.3, 232.4, 231.3, 232.4, 236.4, 258.4, 253.7, 227.8, 220.3, 214.5, 206.5]
JP_HH_LEVERAGE = [62.1, 61.1, 60.0, 60.2, 63.5, 61.2, 61.8, 61.4, 61.3, 60.7, 59.5, 59.8, 60.2, 61.2, 62.3, 67.2, 64.8, 64.7, 62.4, 62.1, 61.1]
JP_CORP_LEVERAGE = [310.9, 307.3, 303.9, 313.9, 339.3, 337.7, 351.0, 356.5, 360.6, 367.2, 360.0, 368.1, 368.2, 373.0, 382.9, 422.5, 415.2, 396.0, 384.3, 371.3, 353.9]

# 样式定义
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
data_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_sheet(ws, title, headers, data_rows, col_widths=None):
    """创建格式化的工作表"""
    ws.title = title
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 写入数据
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.font = data_font
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = Font(bold=True, size=10)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 格式化百分比
                if isinstance(value, (int, float)):
                    cell.number_format = '0.0"%"'
    
    # 设置列宽
    if col_widths:
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width


def main():
    print("Creating Excel file...")
    
    wb = Workbook()
    
    # ========== Sheet 1: 完整数据 ==========
    ws1 = wb.active
    headers = ['指标'] + [str(y) for y in YEARS]
    
    data_rows = [
        ['美国政府债务/GDP'] + US_GOV_DEBT,
        ['美国居民杠杆率'] + US_HH_LEVERAGE,
        ['美国企业杠杆率'] + US_CORP_LEVERAGE,
        ['中国政府债务/GDP'] + CN_GOV_DEBT,
        ['中国居民杠杆率'] + CN_HH_LEVERAGE,
        ['中国企业杠杆率'] + CN_CORP_LEVERAGE,
        ['日本政府债务/GDP'] + JP_GOV_DEBT,
        ['日本居民杠杆率'] + JP_HH_LEVERAGE,
        ['日本企业杠杆率'] + JP_CORP_LEVERAGE,
    ]
    
    col_widths = [20] + [8] * 21
    create_sheet(ws1, '完整数据', headers, data_rows, col_widths)
    
    # ========== Sheet 2: 美国 ==========
    ws2 = wb.create_sheet()
    us_headers = ['指标'] + [str(y) for y in YEARS]
    us_data = [
        ['政府债务/GDP'] + US_GOV_DEBT,
        ['居民杠杆率'] + US_HH_LEVERAGE,
        ['企业杠杆率'] + US_CORP_LEVERAGE,
    ]
    create_sheet(ws2, '美国', us_headers, us_data, col_widths)
    
    # ========== Sheet 3: 中国 ==========
    ws3 = wb.create_sheet()
    cn_data = [
        ['政府债务/GDP'] + CN_GOV_DEBT,
        ['居民杠杆率'] + CN_HH_LEVERAGE,
        ['企业杠杆率'] + CN_CORP_LEVERAGE,
    ]
    create_sheet(ws3, '中国', us_headers, cn_data, col_widths)
    
    # ========== Sheet 4: 日本 ==========
    ws4 = wb.create_sheet()
    jp_data = [
        ['政府债务/GDP'] + JP_GOV_DEBT,
        ['居民杠杆率'] + JP_HH_LEVERAGE,
        ['企业杠杆率'] + JP_CORP_LEVERAGE,
    ]
    create_sheet(ws4, '日本', us_headers, jp_data, col_widths)
    
    # 保存
    wb.save(OUT_PATH)
    print(f"✅ Excel saved: {OUT_PATH}")
    print(f"   - Sheet1: 完整数据 (9行×22列)")
    print(f"   - Sheet2: 美国 (3行×22列)")
    print(f"   - Sheet3: 中国 (3行×22列)")
    print(f"   - Sheet4: 日本 (3行×22列)")


if __name__ == "__main__":
    main()
