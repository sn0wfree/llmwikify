"""
将画图所需数据整合到 Excel 文件
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = os.path.expanduser("~/llmwikify/TopicResearch")
DATA_PATH = os.path.join(BASE_DIR, "data/raw/macro_4country_2000_2026.json")
OUT_PATH = os.path.join(BASE_DIR, "report/chart_data.xlsx")

# 加载数据
with open(DATA_PATH) as f:
    data = json.load(f)

years = data['years']

# 创建工作簿
wb = Workbook()

# 样式定义
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_sheet(ws, title, headers, data_rows):
    """创建格式化的工作表"""
    ws.title = title
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='right')
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15


# ========== Sheet 1: 名义GDP增速 ==========
ws1 = wb.active
headers = ['年份', '美国', '中国', '日本', '欧盟']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['gdp_nom']['us'][i],
        data['gdp_nom']['cn'][i],
        data['gdp_nom']['jp'][i],
        data['gdp_nom']['eu'][i],
    ])
create_sheet(ws1, '名义GDP增速', headers, rows)

# ========== Sheet 2: 实际GDP增速 ==========
ws2 = wb.create_sheet()
headers = ['年份', '美国', '中国', '日本', '欧盟']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['gdp_real']['us'][i],
        data['gdp_real']['cn'][i],
        data['gdp_real']['jp'][i],
        data['gdp_real']['eu'][i],
    ])
create_sheet(ws2, '实际GDP增速', headers, rows)

# ========== Sheet 3: CPI通胀 ==========
ws3 = wb.create_sheet()
headers = ['年份', '美国', '中国', '日本', '欧盟']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['cpi']['us'][i],
        data['cpi']['cn'][i],
        data['cpi']['jp'][i],
        data['cpi']['eu'][i],
    ])
create_sheet(ws3, 'CPI通胀', headers, rows)

# ========== Sheet 4: 货币政策 ==========
ws4 = wb.create_sheet()
headers = ['年份', '中国RRR(%)', '中国M2增速(%)', '中国LPR(%)']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['cn_monetary']['rrr'][i],
        data['cn_monetary']['m2'][i],
        data['cn_monetary']['lpr'][i],
    ])
create_sheet(ws4, '货币政策', headers, rows)

# ========== Sheet 5: 政策利率 ==========
ws5 = wb.create_sheet()
headers = ['年份', '美国(%)', '中国(%)', '日本(%)', '欧盟(%)']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['rate']['us'][i],
        data['rate']['cn'][i],
        data['rate']['jp'][i],
        data['rate']['eu'][i],
    ])
create_sheet(ws5, '政策利率', headers, rows)

# ========== Sheet 6: 市场利率 ==========
ws6 = wb.create_sheet()
headers = ['年份', '中国短端', '中国长端', '美国短端', '美国长端', '日本短端', '日本长端']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['market_rate']['cn_short'][i],
        data['market_rate']['cn_long'][i],
        data['market_rate']['us_short'][i],
        data['market_rate']['us_long'][i],
        data['market_rate']['jp_short'][i],
        data['market_rate']['jp_long'][i],
    ])
create_sheet(ws6, '市场利率', headers, rows)

# ========== Sheet 7: 汇率 ==========
ws7 = wb.create_sheet()
headers = ['年份', 'EUR/USD', 'USD/JPY', 'USD/CNY', '美元指数']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['fx']['eurusd'][i],
        data['fx']['usdjpy'][i],
        data['fx']['usdcny'][i],
        data['dxy'][i],
    ])
create_sheet(ws7, '汇率', headers, rows)

# ========== Sheet 8: 贸易差额 ==========
ws8 = wb.create_sheet()
headers = ['年份', '中国', '美国', '欧盟', '日本']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['trade']['cn'][i],
        data['trade']['us'][i],
        data['trade']['eu'][i],
        data['trade']['jp'][i],
    ])
create_sheet(ws8, '贸易差额', headers, rows)

# ========== Sheet 9: 外汇储备 ==========
ws9 = wb.create_sheet()
headers = ['年份', '中国', '美国', '日本']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['forex']['cn'][i],
        data['forex']['us'][i],
        data['forex']['jp'][i],
    ])
create_sheet(ws9, '外汇储备', headers, rows)

# ========== Sheet 10: GDP(USD) ==========
ws10 = wb.create_sheet()
headers = ['年份', '中国', '美国', '日本']
rows = []
for i, y in enumerate(years):
    rows.append([
        int(y),
        data['gdp_usd']['cn'][i],
        data['gdp_usd']['us'][i],
        data['gdp_usd']['jp'][i],
    ])
create_sheet(ws10, 'GDP(USD)', headers, rows)

# 保存
wb.save(OUT_PATH)
print(f"✅ Excel 已保存: {OUT_PATH}")
print(f"   包含 10 个工作表")
